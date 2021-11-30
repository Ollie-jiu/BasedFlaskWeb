from app.controller.createCourse import CourseName
from flask import Blueprint,render_template, request
from app.models.base import db
from app.models.course import Course
from app.controller.forms import NewPostForm
from flask_login import login_required
from app.models.course_cilo import course_cilo
from app.models.assessment import assessment
import random
from openpyxl import load_workbook
modifyCourseBP=Blueprint('modifyCourse',__name__)
modifyCILOs_AssessmentsBP=Blueprint('modifyCILOs_Assessments',__name__)
assessmentList=[]
CILOsList=[]
relationship=[]
coursesInfo=[]
weight=[]
CIlOexcelFile=[]
assessmentexcelFile=[]
cilo=[]
temp_cilo=[]#a temp list to store cilo result
courseID=random.randint(1000,9999)#randomly generate the courseID
courseSelected=[]#存要modify的课程
@modifyCourseBP.route('/modifyCourse?email=<email>', methods=['GET','POST'])
@login_required
def modifyCourse(email):
    
    form=NewPostForm()
    if request.method=="GET":
        courseList=[]
        temp=Course.query.filter().group_by(Course.name).all()
        for i in temp:
            courseList.append(i.name)
        # print(courseList)
        # print('selected:',courseSelected)
        if courseSelected!=[]:
            #通过这个courseSelected去搜索信息
            #前端只需要courseInfo的List；
            temp=Course.query.filter(Course.name==courseSelected[0]).first().jsonstr()
            return render_template('modifyCourse.html',message="",courseList=courseList,coursesInfo=temp,email=email,function="MODIFY A COURSE",Wtype="Course designer")
        else:
        #有一个需要返回回来的那个之前输入的东西；
        #courseList 需要那个所有的课程；让它等于courseList
            #print('selected:',courseSelected)
            return render_template('modifyCourse.html',message="",coursesInfo={},courseList=courseList,email=email,function="MODIFY A COURSE",Wtype="Course designer")
    #POST method
    else:
        #！！！！！！要一直拉去CourseList为了显示下拉框
        courseList=[]
        temp=Course.query.filter().group_by(Course.name).all()
        for i in temp:
            courseList.append(i.name)
        print('relation',relationship)
        if temp_cilo==[]:
            #用于添加relationship之前
            if courseSelected==[]:
                course=request.form.get("courseName")
                courseSelected.append(course)
                #print(courseSelected)
            #通过这个courseSelected去数据找到所有的课程数据（name-type）
            print(courseSelected)
            temp=Course.query.filter(Course.name==courseSelected[0]).first().jsonstr()
            print(temp)
            coursesInfo.append((temp['CourseName'],temp['_type']))
            print(coursesInfo)
            return render_template('modifyCourse.html',message="",courseList=courseList,courseSelected=[],coursesInfo=temp,email=email,function="MODIFY A COURSE",Wtype="Course designer")
        else:
            print('1:',courseSelected)
            #submit所有东西和清空
            #在这里拿到所有的input值
            #需要把button移到最上面
            print(assessmentList)
            for i in range(0,len(assessmentList)):
                ass=assessment((Course.query.filter(Course.name==courseSelected[0]).first().jsonstr())['courseID'], (Course.query.filter(Course.name==courseSelected[0]).first().jsonstr())['semester'],cilo[i],assessmentList[i],weight[i]) #type is get from assessmentList
                db.session.add(ass)
                db.session.commit()
            
            coursesInfo.clear()
            weight.clear()
            cilo.clear()
            courseSelected.clear()
            temp_cilo.clear()
            assessmentList.clear()
            CILOsList.clear()
            relationship.clear()
            return render_template('modifyCourse.html',message="success",courseList=courseList,courseSelected=[],coursesInfo={},email=email,function="MODIFY A COURSE",Wtype="Course designer")
        
@modifyCILOs_AssessmentsBP.route('/modifyCILOs_Assessments?email=<email>', methods=['GET','POST'])
def modifyCILO_Assessment(email):
    form=NewPostForm()
    if request.method == "GET":
        CILO_order=[]
        if len(CILOsList)==2:
            CILO_order=["CILO1","CILO2"]
        elif len(CILOsList)==3:
            CILO_order=["CILO1","CILO2","CILO3"]
        elif len(CILOsList)==1:
            CILO_order=["CILO1"]
        else:
            CILO_order=[]
        return render_template('cilo_assessment.html',success="",Assessmentmessage='',message="",CILO_order=CILO_order,assessmentList=assessmentList,CILO=CILOsList,form=form,email=email,function="Modify CILOS&ASSESSMENTS",Wtype="Course designer")
    else:
        assessmentMethod=''
        if form.validate_on_submit():
            if form.addCILO.data:
                ExcelName=request.form.to_dict()["filename"]
                if ExcelName=="":
                    CILO1_value=request.form.get("CILO")
                    if len(CILOsList)<3:
                        if CILO1_value!='':
                            CILOsList.append(CILO1_value)
                            print(len(CILOsList))
                        message=""
                    else:
                        message="FULL"
                    if len(CILOsList)==2:
                        CILO_order=["CILO1","CILO2"]
                    elif len(CILOsList)==3:
                        CILO_order=["CILO1","CILO2","CILO3"]
                    elif len(CILOsList)==1:
                        CILO_order=["CILO1"]
                    else:
                        CILO_order=[]
                    return render_template('cilo_assessment.html',success="",Assessmentmessage='',Wtype="Course designer",assessmentList=assessmentList,message=message,CILO_order=CILO_order,CILO=CILOsList,form=form,assessmentMethods=assessmentMethod,email=email,function="Modify CILOS&ASSESSMENTS")
                else:
                    print(ExcelName)
                    get_data(ExcelName)
                    CILO1_value=request.form.get("CILO")
                    for i in CIlOexcelFile:
                        print(i)
                        for j in i:
                            CILOsList.append(j)
                    print(CILOsList)
                    if len(CILOsList)<3:
                        if CILO1_value!='':
                            print(len(CILOsList))
                        message=""
                    else:
                        message="FULL"
                    if len(CILOsList)==2:
                        CILO_order=["CILO1","CILO2"]
                    elif len(CILOsList)==3:
                        CILO_order=["CILO1","CILO2","CILO3"]
                    elif len(CILOsList)==1:
                        CILO_order=["CILO1"]
                    else:
                        CILO_order=[]
                    
                    return render_template('cilo_assessment.html',success="",Assessmentmessage="",Wtype="Course designer",assessmentList=assessmentList,message=message,CILO_order=CILO_order,CILO=CILOsList,form=form,assessmentMethods=assessmentMethod,email=email,function="Modify CILOS&ASSESSMENTS")
            elif form.add.data:
                CILO_order=[]
                if len(CILOsList)==2:
                    CILO_order=["CILO1","CILO2"]
                elif len(CILOsList)==3:
                    CILO_order=["CILO1","CILO2","CILO3"]
                elif len(CILOsList)==1:
                    CILO_order=["CILO1"]
                else:
                    CILO_order=[]
                ExcelName=request.form.to_dict()["assessmentFile"]
                if ExcelName=="":
                    assessmentMethod=request.form.get("AssessmentMethod")
                    if assessmentMethod=="":
                        return render_template('cilo_assessment.html',success="",Assessmentmessage="none",message='',Wtype="Course designer",CILO=CILOsList,CILO_order=CILO_order,form=form,assessmentList=assessmentList,email=email,function="Modify CILOS&ASSESSMENTS")
                    else:
                        assessmentList.append(assessmentMethod)
                        return render_template('cilo_assessment.html',success="",Assessmentmessage='',message="",Wtype="Course designer",CILO=CILOsList,CILO_order=CILO_order,form=form,assessmentList=assessmentList,email=email,function="Modify CILOS&ASSESSMENTS")
                else:
                    get_data(ExcelName)
                    for i in assessmentexcelFile:
                        print(i)
                        for j in i:
                            assessmentList.append(j)
                    if assessmentMethod==""or assessmentList!="":
                        return render_template('cilo_assessment.html',success="",Assessmentmessage="none",message='',Wtype="Course designer",CILO=CILOsList,CILO_order=CILO_order,form=form,assessmentList=assessmentList,email=email,function="Modify CILOS&ASSESSMENTS")
                    else:
                        assessmentList.append(assessmentMethod)
                        return render_template('cilo_assessment.html',success="",Assessmentmessage='',message="",Wtype="Course designer",CILO=CILOsList,CILO_order=CILO_order,form=form,assessmentList=assessmentList,email=email,function="Modify CILOS&ASSESSMENTS")
            else:
                
                relationship=request.form
                temp=dict(relationship.lists())#change to list convinient for getting data
                print(temp)
                try:
                    for i in assessmentList:
                        try:
                            weight.append(int(relationship['percentage:'+i]))
                        except ValueError:
                            print("do not input same string")
                        temp_cilo.append(temp[i])
                except KeyError:
                    pass
                #print('relation:',relationship)    
                #能拿值，但是当勾了cilo1和cilo3的时候， 会添加成1-3; 选了1，2，3时会添加1-2-3
                for i in temp_cilo:
                    result=''
                    counter=0
                    for j in i:
                        if counter!=len(i)-1:
                            if j[4] not in result:
                                result+=j[4]+','
                        else:
                            if j[4] not in result:
                                result+=j[4]
                        counter+=1
                    cilo.append(result)
                print(CILOsList)
                print('cilo:',cilo)
                return render_template('cilo_assessment.html',success="success",Assessmentmessage='',Wtype="Course designer",message="",CILO="",CILO_order="",form=form,assessmentList="",email=email,function="Modify CILOS&ASSESSMENTS")
def get_data(fileName):
    if fileName=="Assessments.xlsx":
        print("....")
        e = load_workbook('C:\\Users\\50627\\Desktop\Balsam_project\\code\\project_Balsam\\app\\static\\excel\\Assessments.xlsx')
        sheet = e["Sheet1"]
        test_data = []  #数据存储为列表格式
        for i in range(1,sheet.max_row):
            sub_data = {} #数据存储为字典格式
            sub_data['method1'] = sheet.cell(i+1,1).value
            sub_data['method2'] = sheet.cell(i+1,2).value
            sub_data['method3'] = sheet.cell(i+1,3).value
            sub_data['method4'] = sheet.cell(i+1,4).value
        assessmentexcelFile.append(list(sub_data.values()))
    if fileName=="CILO.xlsx":
        print("....")
        e = load_workbook('C:\\Users\\50627\\Desktop\\Balsam_project\\code\\project_Balsam\\app\\static\\excel\\CILO.xlsx')
        sheet = e["Sheet1"]
        test_data = []  #数据存储为列表格式
        for i in range(1,sheet.max_row):
            sub_data = {} #数据存储为字典格式
            sub_data['cilo1'] = sheet.cell(i+1,1).value
            sub_data['cilo2'] = sheet.cell(i+1,2).value
        CIlOexcelFile.append(list(sub_data.values()))
    return 1