from flask import Blueprint, json,render_template, request
from app.models.base import db
from app.models.course import Course
from flask_login import login_required
from app.controller.forms import NewPostForm
from app.models.assessment import assessment
from werkzeug.datastructures import ImmutableMultiDict
from openpyxl import load_workbook
import random
from app.models.course_cilo import course_cilo
createCourseBP=Blueprint('createCourse',__name__)
addCILO_AssessmentBP=Blueprint('addCILO_Assessment',__name__)
assessmentList=[]
CILOsList=[]
relationship=[]
CourseName=''
CourseCode=''
semester=''
programme=''
type=''
weight=[]
CIlOexcelFile=[]
assessmentexcelFile=[]
cilo=[]
temp_cilo=[]#a temp list to store cilo result
courseID=random.randint(1000,9999)#randomly generate the courseID
@createCourseBP.route('/createCourse?email=<email>', methods=['GET','POST'])
@login_required
def createCourse(email):
    form=NewPostForm()
    if request.method == "GET":
        return render_template('createCourse.html',form=form,email=email,function="CREATE A COURSE",Wtype="Course designer")
    #post method
    else:
        if form.validate_on_submit():
            CILO_order=[]
            if len(CILOsList)==2:
                CILO_order=["CILO1","CILO2"]
            elif len(CILOsList)==3:
                CILO_order=["CILO1","CILO2","CILO3"]
            elif len(CILOsList)==1:
                CILO_order=["CILO1"]
            else:
                CILO_order=[]
            
            CourseName=request.form.get("CourseName")
            CourseCode=request.form.get("CourseCode")
            semester=request.form.get("semester")
            precourseID=0#default precourseID
            programme=request.form.get("programme")
            type=request.form.get("type")
            #add course
            course=Course(courseID, CourseName, CourseCode, semester,precourseID,programme,type)
            db.session.add(course)
            db.session.commit()
            #add assessment
            print('a',assessmentList)
            for i in range(0,len(assessmentList)):
                ass=assessment(courseID,semester,cilo[i],assessmentList[i],weight[i]) #type is get from assessmentList
                db.session.add(ass)
                db.session.commit()
            assessmentexcelFile.clear()
            CIlOexcelFile.clear()
            counter=0
            for i in CILOsList:
                counter+=1
                cour_cilo=course_cilo(courseID,str(courseID)+'-'+'cilo'+str(counter),i)
                db.session.add(cour_cilo)
                db.session.commit()
                print(1)
            assessmentList.clear()
            CILOsList.clear()
            relationship.clear()
            cilo.clear()
            weight.clear()
            temp_cilo.clear()#a temp list to store cilo result
            return render_template('createCourse.html',message="",CILO_order=CILO_order,assessmentList=assessmentList,Wtype="Course designer",CILO=CILOsList,form=form,email=email,function="ADD CILOS&ASSESSMENTS") 
@addCILO_AssessmentBP.route('/addCILO_Assessment?email=<email>', methods=['GET','POST'])
def addCILO_Assessment(email):
    form=NewPostForm()
    if request.method == "GET":
        CILO_order=[]
        if len(CILOsList)==2:
            CILO_order=["CILO1","CILO2"]
        elif len(CILOsList)==3:
            CILO_order=["CILO1","CILO2","CILO3"]
        elif len(CILOsList)==1:
            CILO_order=["CILO1"]
        else:
            CILO_order=[]
        return render_template('cilo_assessment.html',success="",Assessmentmessage='',message="",CILO_order=CILO_order,assessmentList=assessmentList,CILO=CILOsList,form=form,email=email,function="ADD CILOS&ASSESSMENTS",Wtype="Course designer")
    #post method
    else:
        assessmentMethod=''
        if form.validate_on_submit():
            if form.addCILO.data:
                ExcelName=request.form.to_dict()["filename"]
                if ExcelName=="":
                    CILO1_value=request.form.get("CILO")
                    if len(CILOsList)<3:
                        if CILO1_value!='':
                            CILOsList.append(CILO1_value)
                            print(len(CILOsList))
                        message=""
                    else:
                        message="FULL"
                    if len(CILOsList)==2:
                        CILO_order=["CILO1","CILO2"]
                    elif len(CILOsList)==3:
                        CILO_order=["CILO1","CILO2","CILO3"]
                    elif len(CILOsList)==1:
                        CILO_order=["CILO1"]
                    else:
                        CILO_order=[]
                    return render_template('cilo_assessment.html',success="",Assessmentmessage='',Wtype="Course designer",assessmentList=assessmentList,message=message,CILO_order=CILO_order,CILO=CILOsList,form=form,assessmentMethods=assessmentMethod,email=email,function="ADD CILOS&ASSESSMENTS")
                else:
                    print(ExcelName)
                    get_data(ExcelName)
                    CILO1_value=request.form.get("CILO")
                    for i in CIlOexcelFile:
                        print(i)
                        for j in i:
                            CILOsList.append(j)
                    print(CILOsList)
                    if len(CILOsList)<3:
                        if CILO1_value!='':
                            print(len(CILOsList))
                        message=""
                    else:
                        message="FULL"
                    if len(CILOsList)==2:
                        CILO_order=["CILO1","CILO2"]
                    elif len(CILOsList)==3:
                        CILO_order=["CILO1","CILO2","CILO3"]
                    elif len(CILOsList)==1:
                        CILO_order=["CILO1"]
                    else:
                        CILO_order=[]
                    
                    return render_template('cilo_assessment.html',success="",Assessmentmessage="",Wtype="Course designer",assessmentList=assessmentList,message=message,CILO_order=CILO_order,CILO=CILOsList,form=form,assessmentMethods=assessmentMethod,email=email,function="ADD CILOS&ASSESSMENTS")
            elif form.add.data:
                CILO_order=[]
                if len(CILOsList)==2:
                    CILO_order=["CILO1","CILO2"]
                elif len(CILOsList)==3:
                    CILO_order=["CILO1","CILO2","CILO3"]
                elif len(CILOsList)==1:
                    CILO_order=["CILO1"]
                else:
                    CILO_order=[]
                ExcelName=request.form.to_dict()["assessmentFile"]
                if ExcelName=="":
                    assessmentMethod=request.form.get("AssessmentMethod")
                    if assessmentMethod=="":
                        return render_template('cilo_assessment.html',success="",Assessmentmessage="none",message='',Wtype="Course designer",CILO=CILOsList,CILO_order=CILO_order,form=form,assessmentList=assessmentList,email=email,function="ADD CILOS&ASSESSMENTS")
                    else:
                        assessmentList.append(assessmentMethod)
                        return render_template('cilo_assessment.html',success="",Assessmentmessage='',message="",Wtype="Course designer",CILO=CILOsList,CILO_order=CILO_order,form=form,assessmentList=assessmentList,email=email,function="ADD CILOS&ASSESSMENTS")
                else:
                    get_data(ExcelName)
                    for i in assessmentexcelFile:
                        print(i)
                        for j in i:
                            assessmentList.append(j)
                    if assessmentMethod==""or assessmentList!="":
                        return render_template('cilo_assessment.html',success="",Assessmentmessage="none",message='',Wtype="Course designer",CILO=CILOsList,CILO_order=CILO_order,form=form,assessmentList=assessmentList,email=email,function="ADD CILOS&ASSESSMENTS")
                    else:
                        assessmentList.append(assessmentMethod)
                        return render_template('cilo_assessment.html',success="",Assessmentmessage='',message="",Wtype="Course designer",CILO=CILOsList,CILO_order=CILO_order,form=form,assessmentList=assessmentList,email=email,function="ADD CILOS&ASSESSMENTS")
            else:
                
                relationship=request.form
                temp=dict(relationship.lists())#change to list convinient for getting data
                print(temp)
                try:
                    for i in assessmentList:
                        try:
                            weight.append(int(relationship['percentage:'+i]))
                        except ValueError:
                            print("do not input same string")
                        temp_cilo.append(temp[i])
                except KeyError:
                    pass
                #print('relation:',relationship)    
                #能拿值，但是当勾了cilo1和cilo3的时候， 会添加成1-3; 选了1，2，3时会添加1-2-3
                for i in temp_cilo:
                    result=''
                    counter=0
                    for j in i:
                        if counter!=len(i)-1:
                            if j[4] not in result:
                                result+=j[4]+','
                        else:
                            if j[4] not in result:
                                result+=j[4]
                        counter+=1
                    cilo.append(result)
                print(CILOsList)
                print('cilo:',cilo)
                return render_template('cilo_assessment.html',success="success",Assessmentmessage='',Wtype="Course designer",message="",CILO="",CILO_order="",form=form,assessmentList="",email=email,function="ADD CILOS&ASSESSMENTS")
def get_data(fileName):
    if fileName=="Assessments.xlsx":
        print("....")
        e = load_workbook('C:\\Users\\50627\\Desktop\Balsam_project\\code\\project_Balsam\\app\\static\\excel\\Assessments.xlsx')
        sheet = e["Sheet1"]
        test_data = []  #数据存储为列表格式
        for i in range(1,sheet.max_row):
            sub_data = {} #数据存储为字典格式
            sub_data['method1'] = sheet.cell(i+1,1).value
            sub_data['method2'] = sheet.cell(i+1,2).value
            sub_data['method3'] = sheet.cell(i+1,3).value
            sub_data['method4'] = sheet.cell(i+1,4).value
        assessmentexcelFile.append(list(sub_data.values()))
    if fileName=="CILO.xlsx":
        print("....")
        e = load_workbook('C:\\Users\\50627\\Desktop\\Balsam_project\\code\\project_Balsam\\app\\static\\excel\\CILO.xlsx')
        sheet = e["Sheet1"]
        test_data = []  #数据存储为列表格式
        for i in range(1,sheet.max_row):
            sub_data = {} #数据存储为字典格式
            sub_data['cilo1'] = sheet.cell(i+1,1).value
            sub_data['cilo2'] = sheet.cell(i+1,2).value
        CIlOexcelFile.append(list(sub_data.values()))
    return 1